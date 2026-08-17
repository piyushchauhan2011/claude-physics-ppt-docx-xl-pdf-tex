#!/usr/bin/env python3
"""Generate 'Projectile_Motion.xlsx' — 2D projectile motion workbook with
an embedded trajectory chart, built entirely with live Excel formulas."""

import math

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import ScatterChart, Series, Reference
from openpyxl.chart.axis import ChartLines

# ---------------------------------------------------------------------------
# Default input values (used only to size the trajectory table; every cell
# in the sheet recalculates live from B3:B6 if the user edits them)
# ---------------------------------------------------------------------------
V0 = 25.0
THETA = 45.0
G = 9.81
DT = 0.1

vy0 = V0 * math.sin(math.radians(THETA))
t_total = 2 * vy0 / G
n_steps = math.floor(t_total / DT)  # number of dt-sized steps after t=0

# ---------------------------------------------------------------------------
# Styling constants
# ---------------------------------------------------------------------------
NAVY = "0B1F3A"
TEAL = "2EC4B6"
TEAL_DARK = "1E8F84"
LIGHT_BLUE = "EAF3F7"
LIGHT_GREEN = "EAF7EE"
HEADER_GRAY = "E9EDF3"
WHITE = "FFFFFF"
TEXT_DARK = "1B222D"

FONT_NAME = "Calibri"

title_font = Font(name=FONT_NAME, size=18, bold=True, color=WHITE)
section_font = Font(name=FONT_NAME, size=12, bold=True, color=WHITE)
label_font = Font(name=FONT_NAME, size=11, color=TEXT_DARK)
value_font = Font(name=FONT_NAME, size=11, bold=True, color=NAVY)
table_header_font = Font(name=FONT_NAME, size=11, bold=True, color=WHITE)
table_body_font = Font(name=FONT_NAME, size=10.5, color=TEXT_DARK)
note_font = Font(name=FONT_NAME, size=9, italic=True, color="6B7484")

title_fill = PatternFill("solid", fgColor=NAVY)
section_fill = PatternFill("solid", fgColor=TEAL_DARK)
input_fill = PatternFill("solid", fgColor=LIGHT_BLUE)
summary_fill = PatternFill("solid", fgColor=LIGHT_GREEN)
table_header_fill = PatternFill("solid", fgColor=NAVY)
table_alt_fill = PatternFill("solid", fgColor=HEADER_GRAY)

thin = Side(style="thin", color="C7CFDB")
box_border = Border(left=thin, right=thin, top=thin, bottom=thin)

center = Alignment(horizontal="center", vertical="center")
left = Alignment(horizontal="left", vertical="center")
right = Alignment(horizontal="right", vertical="center")

# ---------------------------------------------------------------------------
# Workbook / sheet
# ---------------------------------------------------------------------------
wb = Workbook()
ws = wb.active
ws.title = "Projectile Motion"
ws.sheet_view.showGridLines = False

for col, w in zip("ABCDEFGH", [30, 16, 16, 16, 16, 4, 4, 4]):
    ws.column_dimensions[col].width = w

# ---------------------------------------------------------------------------
# Title
# ---------------------------------------------------------------------------
ws.merge_cells("A1:E1")
ws["A1"] = "2D Projectile Motion — Trajectory Simulator"
ws["A1"].font = title_font
ws["A1"].fill = title_fill
ws["A1"].alignment = left
ws.row_dimensions[1].height = 30
for col in "ABCDE":
    ws[f"{col}1"].fill = title_fill

# ---------------------------------------------------------------------------
# Section 1 — Input Parameters (rows 2-6; v0 -> B3, theta -> B4, g -> B5, dt -> B6)
# ---------------------------------------------------------------------------
ws.merge_cells("A2:B2")
ws["A2"] = "1.  INPUT PARAMETERS"
ws["A2"].font = section_font
ws["A2"].fill = section_fill
for col in "AB":
    ws[f"{col}2"].fill = section_fill
ws.row_dimensions[2].height = 20

inputs = [
    (3, "Initial Velocity, v₀ (m/s)", V0, "0.00"),
    (4, "Launch Angle, θ (degrees)", THETA, "0.0"),
    (5, "Acceleration due to Gravity, g (m/s²)", G, "0.00"),
    (6, "Time Step, dt (s)", DT, "0.00"),
]
for row, label, val, fmt in inputs:
    ws[f"A{row}"] = label
    ws[f"A{row}"].font = label_font
    ws[f"A{row}"].alignment = left
    ws[f"B{row}"] = val
    ws[f"B{row}"].font = value_font
    ws[f"B{row}"].fill = input_fill
    ws[f"B{row}"].number_format = fmt
    ws[f"B{row}"].alignment = center
    ws[f"B{row}"].border = box_border
    ws[f"A{row}"].border = box_border

# ---------------------------------------------------------------------------
# Section 2 — Analytical Summary (rows 7-13)
# ---------------------------------------------------------------------------
ws.merge_cells("A8:B8")
ws["A8"] = "2.  ANALYTICAL SUMMARY"
ws["A8"].font = section_font
ws["A8"].fill = section_fill
for col in "AB":
    ws[f"{col}8"].fill = section_fill
ws.row_dimensions[8].height = 20

summary = [
    (9, "Initial Horizontal Velocity, Vx (m/s)", "=B3*COS(RADIANS(B4))", "0.000"),
    (10, "Initial Vertical Velocity, Vy (m/s)", "=B3*SIN(RADIANS(B4))", "0.000"),
    (11, "Total Time of Flight, t_total (s)", "=2*B10/B5", "0.000"),
    (12, "Maximum Height, H_max (m)", "=B10^2/(2*B5)", "0.000"),
    (13, "Total Range, R_max (m)", "=B3^2*SIN(RADIANS(2*B4))/B5", "0.000"),
]
for row, label, formula, fmt in summary:
    ws[f"A{row}"] = label
    ws[f"A{row}"].font = label_font
    ws[f"A{row}"].alignment = left
    ws[f"B{row}"] = formula
    ws[f"B{row}"].font = value_font
    ws[f"B{row}"].fill = summary_fill
    ws[f"B{row}"].number_format = fmt
    ws[f"B{row}"].alignment = center
    ws[f"B{row}"].border = box_border
    ws[f"A{row}"].border = box_border

# ---------------------------------------------------------------------------
# Section 3 — Trajectory Data Table (header row 15, data from row 16)
# ---------------------------------------------------------------------------
ws.merge_cells("A15:E15")
ws["A15"] = "3.  TRAJECTORY DATA TABLE"
ws["A15"].font = section_font
ws["A15"].fill = section_fill
for col in "ABCDE":
    ws[f"{col}15"].fill = section_fill
ws.row_dimensions[15].height = 20

headers = ["Time t (s)", "Distance x (m)", "Height y (m)", "Vx (m/s)", "Vy (m/s)"]
HEADER_ROW = 16
for i, h in enumerate(headers):
    col = get_column_letter(1 + i)
    cell = ws[f"{col}{HEADER_ROW}"]
    cell.value = h
    cell.font = table_header_font
    cell.fill = table_header_fill
    cell.alignment = center
    cell.border = box_border

DATA_START = HEADER_ROW + 1  # row 17
total_data_rows = n_steps + 2  # t=0 ... n_steps*dt, plus one exact-landing row
DATA_END = DATA_START + total_data_rows - 1

for i in range(total_data_rows):
    r = DATA_START + i
    is_last = (i == total_data_rows - 1)

    if i == 0:
        ws[f"A{r}"] = 0
    elif is_last:
        ws[f"A{r}"] = "=$B$11"
    else:
        ws[f"A{r}"] = f"=A{r - 1}+$B$6"

    ws[f"B{r}"] = f"=$B$3*COS(RADIANS($B$4))*A{r}"
    ws[f"C{r}"] = f"=MAX(0,($B$3*SIN(RADIANS($B$4))*A{r})-(0.5*$B$5*A{r}^2))"
    ws[f"D{r}"] = "=$B$9"
    ws[f"E{r}"] = f"=$B$10-$B$5*A{r}"

    fill = table_alt_fill if (i % 2 == 1) else None
    for col, fmt in zip("ABCDE", ["0.00", "0.000", "0.000", "0.000", "0.000"]):
        cell = ws[f"{col}{r}"]
        cell.font = table_body_font
        cell.number_format = fmt
        cell.alignment = center
        cell.border = box_border
        if fill:
            cell.fill = fill

ws[f"A{DATA_END + 1}"] = "Note: table rows are sized for the default inputs above; edit v0/theta/g/dt and the values recalculate, though extending far beyond the defaults may need extra rows copied down."
ws[f"A{DATA_END + 1}"].font = note_font
ws.merge_cells(f"A{DATA_END + 1}:E{DATA_END + 1}")
ws.row_dimensions[DATA_END + 1].height = 26
ws[f"A{DATA_END + 1}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

# freeze header row of the table for easier scrolling
ws.freeze_panes = f"A{DATA_START}"

# ---------------------------------------------------------------------------
# Section 4 — Trajectory Chart
# ---------------------------------------------------------------------------
chart = ScatterChart()
chart.title = "Projectile Trajectory (Height vs Distance)"
chart.style = 13
chart.x_axis.title = "Distance x (m)"
chart.y_axis.title = "Height y (m)"
chart.x_axis.majorGridlines = ChartLines()
chart.y_axis.majorGridlines = ChartLines()
chart.x_axis.delete = False
chart.y_axis.delete = False
chart.legend.position = "b"
chart.height = 10
chart.width = 22

x_ref = Reference(ws, min_col=2, min_row=DATA_START, max_row=DATA_END)   # Distance x
y_ref = Reference(ws, min_col=3, min_row=DATA_START, max_row=DATA_END)   # Height y

series = Series(y_ref, x_ref, title="Trajectory (y vs x)")
series.marker.symbol = "circle"
series.marker.size = 4
series.graphicalProperties.line.solidFill = TEAL_DARK
series.graphicalProperties.line.width = 22000  # EMU (~1.75pt)
series.smooth = True

chart.series.append(series)

ws.add_chart(chart, f"G{HEADER_ROW}")

# ---------------------------------------------------------------------------
wb.save("Projectile_Motion.xlsx")
print(f"Saved Projectile_Motion.xlsx  |  trajectory rows {DATA_START}-{DATA_END} ({total_data_rows} rows)")
