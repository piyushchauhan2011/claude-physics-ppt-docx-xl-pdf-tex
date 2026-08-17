#!/usr/bin/env python3
"""Generate 'Projectile_Motion_With_Drag.xlsx' — Euler-method projectile
simulation with quadratic air drag, compared against the ideal (vacuum)
trajectory, with a live-formula workbook and an embedded comparison chart."""

import math

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import ScatterChart, Series, Reference
from openpyxl.chart.axis import ChartLines

# ---------------------------------------------------------------------------
# Default inputs (only used here to size the table; every cell recalculates
# live from the input block if the user edits it)
# ---------------------------------------------------------------------------
V0, THETA, G, DT = 25.0, 45.0, 9.81, 0.02
RHO, CD, R, M = 1.225, 0.47, 0.05, 0.25

A_coeff = math.pi * R ** 2
K = 0.5 * RHO * CD * A_coeff
t_total_ideal = 2 * V0 * math.sin(math.radians(THETA)) / G

# --- Euler-simulate the drag trajectory (mirrors the Excel formulas
#     exactly) purely to find how many rows are needed for the table.
vx = V0 * math.cos(math.radians(THETA))
vy = V0 * math.sin(math.radians(THETA))
x = y = t = 0.0
vtot = math.sqrt(vx ** 2 + vy ** 2)
ax = -(K / M) * vx * vtot
ay = -G - (K / M) * vy * vtot
drag_landed_step = None
step = 0
steps_needed_for_ideal = math.ceil(t_total_ideal / DT)
while True:
    t_new = t + DT
    vx_new = vx + ax * DT
    vy_new = vy + ay * DT
    x_new = x + vx * DT
    y_new = max(0.0, y + vy * DT)
    vtot_new = math.sqrt(vx_new ** 2 + vy_new ** 2)
    if y_new > 0:
        ax_new = -(K / M) * vx_new * vtot_new
        ay_new = -G - (K / M) * vy_new * vtot_new
    else:
        ax_new = 0.0
        ay_new = 0.0
    step += 1
    if y_new == 0.0 and drag_landed_step is None:
        drag_landed_step = step
    t, x, y, vx, vy, ax, ay = t_new, x_new, y_new, vx_new, vy_new, ax_new, ay_new
    if drag_landed_step is not None and step >= steps_needed_for_ideal:
        break
    if step > 20000:
        raise RuntimeError("simulation did not terminate")

TOTAL_STEPS = step  # number of dt-steps after t=0
TOTAL_ROWS = TOTAL_STEPS + 1  # includes row 0

# ---------------------------------------------------------------------------
# Styling constants
# ---------------------------------------------------------------------------
NAVY = "0B1F3A"
TEAL_DARK = "1E8F84"
LIGHT_BLUE = "EAF3F7"
LIGHT_GREEN = "EAF7EE"
HEADER_GRAY = "E9EDF3"
WHITE = "FFFFFF"
TEXT_DARK = "1B222D"
DRAG_COLOR = "E0473A"   # red-orange for the drag series
IDEAL_COLOR = "2E8FC4"  # blue for the ideal series

FONT_NAME = "Calibri"

title_font = Font(name=FONT_NAME, size=18, bold=True, color=WHITE)
section_font = Font(name=FONT_NAME, size=12, bold=True, color=WHITE)
label_font = Font(name=FONT_NAME, size=11, color=TEXT_DARK)
value_font = Font(name=FONT_NAME, size=11, bold=True, color=NAVY)
table_header_font = Font(name=FONT_NAME, size=10.5, bold=True, color=WHITE)
table_body_font = Font(name=FONT_NAME, size=10, color=TEXT_DARK)
note_font = Font(name=FONT_NAME, size=9, italic=True, color="6B7484")

title_fill = PatternFill("solid", fgColor=NAVY)
section_fill = PatternFill("solid", fgColor=TEAL_DARK)
input_fill = PatternFill("solid", fgColor=LIGHT_BLUE)
derived_fill = PatternFill("solid", fgColor=LIGHT_GREEN)
table_header_fill = PatternFill("solid", fgColor=NAVY)
table_alt_fill = PatternFill("solid", fgColor=HEADER_GRAY)

thin = Side(style="thin", color="C7CFDB")
box_border = Border(left=thin, right=thin, top=thin, bottom=thin)

center = Alignment(horizontal="center", vertical="center")
left = Alignment(horizontal="left", vertical="center")

# ---------------------------------------------------------------------------
# Workbook / sheet
# ---------------------------------------------------------------------------
wb = Workbook()
ws = wb.active
ws.title = "Projectile w Drag"
ws.sheet_view.showGridLines = False

widths = {"A": 30, "B": 12, "C": 12, "D": 12, "E": 12, "F": 12,
          "G": 12, "H": 12, "I": 12, "J": 12}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

# ---------------------------------------------------------------------------
# Title
# ---------------------------------------------------------------------------
ws.merge_cells("A1:J1")
ws["A1"] = "2D Projectile Motion — Ideal vs. Quadratic Air Resistance"
ws["A1"].font = title_font
ws["A1"].alignment = left
ws.row_dimensions[1].height = 30
for i in range(1, 11):
    ws.cell(row=1, column=i).fill = title_fill

# ---------------------------------------------------------------------------
# Section 1 — Input Parameters (rows 3-10 -> B3:B10)
# ---------------------------------------------------------------------------
ws.merge_cells("A2:B2")
ws["A2"] = "1.  INPUT PARAMETERS"
ws["A2"].font = section_font
for col in "AB":
    ws[f"{col}2"].fill = section_fill
ws.row_dimensions[2].height = 20

inputs = [
    (3, "Initial Velocity, v₀ (m/s)", V0, "0.00"),
    (4, "Launch Angle, θ (degrees)", THETA, "0.0"),
    (5, "Acceleration due to Gravity, g (m/s²)", G, "0.00"),
    (6, "Time Step, dt (s)", DT, "0.000"),
    (7, "Air Density, ρ (kg/m³)", RHO, "0.000"),
    (8, "Drag Coefficient, Cd (sphere)", CD, "0.00"),
    (9, "Projectile Radius, r (m)", R, "0.000"),
    (10, "Projectile Mass, m (kg)", M, "0.00"),
]
for row, label, val, fmt in inputs:
    ws[f"A{row}"] = label
    ws[f"A{row}"].font = label_font
    ws[f"A{row}"].alignment = left
    ws[f"A{row}"].border = box_border
    ws[f"B{row}"] = val
    ws[f"B{row}"].font = value_font
    ws[f"B{row}"].fill = input_fill
    ws[f"B{row}"].number_format = fmt
    ws[f"B{row}"].alignment = center
    ws[f"B{row}"].border = box_border

# Derived quantities (rows 11-12)
derived = [
    (11, "Cross-sectional Area, A = πr² (m²)", "=PI()*B9^2", "0.000000"),
    (12, "Drag Constant, k = 0.5·ρ·Cd·A", "=0.5*B7*B8*B11", "0.000000"),
]
for row, label, formula, fmt in derived:
    ws[f"A{row}"] = label
    ws[f"A{row}"].font = label_font
    ws[f"A{row}"].alignment = left
    ws[f"A{row}"].border = box_border
    ws[f"B{row}"] = formula
    ws[f"B{row}"].font = value_font
    ws[f"B{row}"].fill = derived_fill
    ws[f"B{row}"].number_format = fmt
    ws[f"B{row}"].alignment = center
    ws[f"B{row}"].border = box_border

# ---------------------------------------------------------------------------
# Section 2 — Numerical Integration Table (Euler's Method)
# ---------------------------------------------------------------------------
ws.merge_cells("A14:J14")
ws["A14"] = "2.  NUMERICAL INTEGRATION TABLE (EULER'S METHOD)"
ws["A14"].font = section_font
for i in range(1, 11):
    ws.cell(row=14, column=i).fill = section_fill
ws.row_dimensions[14].height = 20

headers = ["Time t (s)", "x_drag (m)", "y_drag (m)", "vx (m/s)", "vy (m/s)",
           "v_total (m/s)", "ax (m/s²)", "ay (m/s²)", "x_ideal (m)", "y_ideal (m)"]
HEADER_ROW = 15
for i, h in enumerate(headers):
    col = get_column_letter(1 + i)
    cell = ws[f"{col}{HEADER_ROW}"]
    cell.value = h
    cell.font = table_header_font
    cell.fill = table_header_fill
    cell.alignment = center
    cell.border = box_border

DATA_START = HEADER_ROW + 1
DATA_END = DATA_START + TOTAL_ROWS - 1

for i in range(TOTAL_ROWS):
    r = DATA_START + i
    if i == 0:
        # t = 0 baseline row
        ws[f"A{r}"] = 0
        ws[f"D{r}"] = "=$B$3*COS(RADIANS($B$4))"                       # vx
        ws[f"E{r}"] = "=$B$3*SIN(RADIANS($B$4))"                       # vy
        ws[f"B{r}"] = 0                                                # x_drag
        ws[f"C{r}"] = 0                                                # y_drag
        ws[f"F{r}"] = f"=SQRT(D{r}^2+E{r}^2)"                          # v_total
        ws[f"G{r}"] = f"=-($B$12/$B$10)*D{r}*F{r}"                     # ax
        ws[f"H{r}"] = f"=-$B$5-($B$12/$B$10)*E{r}*F{r}"                # ay
        ws[f"I{r}"] = 0                                                # x_ideal
        ws[f"J{r}"] = 0                                                # y_ideal
    else:
        p = r - 1  # previous row
        ws[f"A{r}"] = f"=A{p}+$B$6"
        ws[f"D{r}"] = f"=D{p}+G{p}*$B$6"                                # vx
        ws[f"E{r}"] = f"=E{p}+H{p}*$B$6"                                # vy
        ws[f"B{r}"] = f"=B{p}+D{p}*$B$6"                                # x_drag
        ws[f"C{r}"] = f"=MAX(0,C{p}+E{p}*$B$6)"                         # y_drag
        ws[f"F{r}"] = f"=SQRT(D{r}^2+E{r}^2)"                           # v_total
        ws[f"G{r}"] = f"=IF(C{r}>0,-($B$12/$B$10)*D{r}*F{r},0)"         # ax
        ws[f"H{r}"] = f"=IF(C{r}>0,-$B$5-($B$12/$B$10)*E{r}*F{r},0)"    # ay
        ws[f"I{r}"] = f"=$B$3*COS(RADIANS($B$4))*A{r}"                  # x_ideal
        ws[f"J{r}"] = f"=MAX(0,($B$3*SIN(RADIANS($B$4))*A{r})-(0.5*$B$5*A{r}^2))"  # y_ideal

    fmt_map = {"A": "0.000", "B": "0.000", "C": "0.000", "D": "0.000", "E": "0.000",
               "F": "0.000", "G": "0.000", "H": "0.000", "I": "0.000", "J": "0.000"}
    fill = table_alt_fill if (i % 2 == 1) else None
    for col in "ABCDEFGHIJ":
        cell = ws[f"{col}{r}"]
        cell.font = table_body_font
        cell.number_format = fmt_map[col]
        cell.alignment = center
        cell.border = box_border
        if fill:
            cell.fill = fill

note_row = DATA_END + 1
ws[f"A{note_row}"] = (
    "Note: drag trajectory reaches the ground and holds y_drag = 0 (ax = ay = 0, "
    "vx/vy frozen) for any remaining rows so the ideal curve has room to complete its arc."
)
ws[f"A{note_row}"].font = note_font
ws.merge_cells(f"A{note_row}:J{note_row}")
ws.row_dimensions[note_row].height = 26
ws[f"A{note_row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

ws.freeze_panes = f"A{DATA_START}"

# ---------------------------------------------------------------------------
# Section 3 — Trajectory Comparison Chart
# ---------------------------------------------------------------------------
chart = ScatterChart()
chart.title = "Projectile Motion: Ideal Vacuum vs. Air Resistance"
chart.style = 13
chart.x_axis.title = "Horizontal Distance (m)"
chart.y_axis.title = "Height (m)"
chart.x_axis.majorGridlines = ChartLines()
chart.y_axis.majorGridlines = ChartLines()
chart.x_axis.delete = False
chart.y_axis.delete = False
chart.legend.position = "b"
chart.height = 11
chart.width = 24

# Series 1: trajectory with drag (x_drag, y_drag)
drag_x = Reference(ws, min_col=2, min_row=DATA_START, max_row=DATA_END)
drag_y = Reference(ws, min_col=3, min_row=DATA_START, max_row=DATA_END)
drag_series = Series(drag_y, drag_x, title="Trajectory with Drag")
drag_series.marker.symbol = "circle"
drag_series.marker.size = 3
drag_series.graphicalProperties.line.solidFill = DRAG_COLOR
drag_series.graphicalProperties.line.width = 22000
drag_series.smooth = True
chart.series.append(drag_series)

# Series 2: ideal (vacuum) trajectory (x_ideal, y_ideal)
ideal_x = Reference(ws, min_col=9, min_row=DATA_START, max_row=DATA_END)
ideal_y = Reference(ws, min_col=10, min_row=DATA_START, max_row=DATA_END)
ideal_series = Series(ideal_y, ideal_x, title="Ideal Trajectory (Vacuum)")
ideal_series.marker.symbol = "circle"
ideal_series.marker.size = 3
ideal_series.graphicalProperties.line.solidFill = IDEAL_COLOR
ideal_series.graphicalProperties.line.width = 22000
ideal_series.graphicalProperties.line.dashStyle = "dash"
ideal_series.smooth = True
chart.series.append(ideal_series)

ws.add_chart(chart, f"L{HEADER_ROW}")

# ---------------------------------------------------------------------------
wb.save("Projectile_Motion_With_Drag.xlsx")
print(f"Saved Projectile_Motion_With_Drag.xlsx | rows {DATA_START}-{DATA_END} "
      f"({TOTAL_ROWS} rows) | drag landed at step {drag_landed_step} "
      f"(t={drag_landed_step*DT:.2f}s), ideal lands ~t={t_total_ideal:.2f}s")
